# Script elaborado por la maestra Perla Marlene Viera Honzalez
# Este script fue modificado por los alumnos: Jairo Santana Garcia,
# Pablo de Jesus Garcia Medina, Axel Manuel Perales Teofilo
# Ernesto Jesus Cano Arenas
from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


def busqueda(organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 25, cuidado!
    resultado = hunter.domain_search(company=organizacion,
                                     limit=1, emails_type='personal')
    return resultado


def guardar_info(data_found, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    hoja["A1"] = "Dominio"
    hoja["B1"] = "Organizacion"
    hoja["C1"] = "País"
    hoja["D1"] = "Estado"
    hoja["E1"] = "Webmails"
    hoja["F1"] = "Emails"

    hoja.cell(2, 1, data_found["domain"])
    hoja.cell(2, 2, data_found["organization"])
    hoja.cell(2, 3, data_found["country"])
    hoja.cell(2, 4, data_found["state"])
    hoja.cell(2, 5, data_found["webmail"])
    hoja.cell(2, 6, data_found["emails"][0]["value"])
    libro.save("Hunter" + organizacion + ".xlsx")

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datos_encontrados = busqueda(orga)
if datos_encontrados is None:
    exit()
else:
    print(datos_encontrados)
    print(type(datos_encontrados))
    guardar_info(datos_encontrados, orga)
